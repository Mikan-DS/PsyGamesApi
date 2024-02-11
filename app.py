import datetime
import io
import json
import os
import typing

from flask import Flask, request, jsonify, abort, render_template, redirect, url_for, send_file, flash
from flask_login import UserMixin, LoginManager, login_user, logout_user, current_user, login_required
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import FlaskForm
from openpyxl.workbook import Workbook
from werkzeug.security import generate_password_hash, check_password_hash
from wtforms.fields.simple import SubmitField, PasswordField
from wtforms.validators import InputRequired


PROJECT_PATH = os.path.dirname(os.path.abspath(__file__))

def get_file(filename: str) -> str:
    return os.path.join(PROJECT_PATH, filename)

def add_traceback():
    import traceback
    mode = "a"
    if not os.path.exists(get_file('logs.txt')):
        mode = "w"

    with open(get_file('logs.txt'), mode) as file:
        file.write("-"*40 + "\n" + traceback.format_exc())

def load_projects():
    projects.clear()
    with open(get_file('projects.json'), "r", encoding="UTF-8") as file:
        projects.update(json.load(file))

config = {}
with open(get_file('api_config.json'), "r", encoding="UTF-8") as file:
    config.update(json.load(file))

app = Flask(__name__)
app.config['SECRET_KEY'] = config["FLASK_SECRET_KEY"]
app.config['SQLALCHEMY_DATABASE_URI'] = config["SQLALCHEMY_DATABASE_URI"]
db = SQLAlchemy(app)
login_manager = LoginManager(app)

projects = {}
load_projects()


class User(UserMixin, db.Model):
    """
    На данном этапе нам необходим лишь один пользователь, но в целях масштабируемости поднимает таблицу
    """
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, index=True)
    password_hash = db.Column(db.String(128))

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class TestResult(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(80), nullable=False)
    name = db.Column(db.String(80), nullable=False)
    ip = db.Column(db.String(80), nullable=False)
    end_time = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    duration = db.Column(db.Integer, nullable=False)
    parameters = db.relationship('TestResultParameter', backref='test_result', lazy=True, cascade="all, delete-orphan")

    def __init__(self, project_name, name, ip, duration, result_parameters):
        self.project_name = project_name
        self.name = name
        self.ip = ip
        self.duration = duration
        self.parameters: typing.Collection[TestResultParameter] = self.create_parameters(result_parameters)

    def create_parameters(self, result_parameters):
        parameters = []
        pairs = result_parameters.split(',')
        for pair in pairs:
            pair = pair.strip()
            if not pair:
                continue
            name, value = pair.split(':')
            parameter = TestResultParameter(self.id, name.strip(), value.strip())
            parameters.append(parameter)
        return parameters

    def __repr__(self):
        return '<TestResult %r>' % self.project_name

    def as_dict(self):
        return {
            'id': self.id,
            'project_name': self.project_name,
            'name': self.name,
            'ip': self.ip,
            'end_time': self.end_time,
            'duration': str(datetime.timedelta(seconds=self.duration)),
            'result_parameters': {
                parameter.name: parameter.value for parameter in self.parameters
            }
        }


class TestResultParameter(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    test_result_id = db.Column(db.Integer, db.ForeignKey('test_result.id', ondelete='CASCADE'), nullable=False)
    name = db.Column(db.String(80), nullable=False)
    value = db.Column(db.String(500), nullable=False)

    def __init__(self, test_result_id, name, value):
        self.test_result_id = test_result_id
        self.name = name
        self.value = value

    def __repr__(self):
        return '<TestResultParameter %r>' % self.name

@app.route('/api/add-result', methods=['POST', "OPTIONS"])
def add_result():

    result = dict(request.args)


    try:
        result["ip"] = request.remote_addr
        result["duration"] = int(result["duration"])

        if result.get("project_name") not in projects.keys():
            abort(400, 'Такого теста не существует')
        ts = TestResult(**result)

        if set(map(lambda x: x.name, ts.parameters)) != set(projects[result.get("project_name")]):
            abort(400, 'Неверные результаты теста')

        db.session.add(ts)
        db.session.commit()

        return jsonify(ts.as_dict())

    except Exception as e:
        import traceback

        abort(400, 'Неверный результат теста ')

class DeleteResultsForm(FlaskForm):
    submit = SubmitField('Удалить')

@app.route('/admin/view-results/<string:project_name>', methods=['GET'])
# @login_required
def psytest_view_results(project_name):

    if not current_user.is_authenticated:
        return redirect(url_for('login'))


    if project_name not in projects.keys():
        abort(404, 'Такой проект не найден!')

    delete_results_form = DeleteResultsForm()

    project_parameters = projects[project_name]
    results = tuple(
        map(lambda x: x.as_dict(),
            TestResult.query.filter(TestResult.project_name == project_name).order_by(TestResult.end_time.desc())
            )
    )
    return render_template('view-test-results.html',
                           project_name=project_name,
                           projects=projects,
                           results=results,
                           form=delete_results_form
                           )

@app.route('/api/results/<string:project_name>/delete', methods=['POST'])
def delete_results(project_name):
    delete_results_form = DeleteResultsForm()

    if delete_results_form.validate_on_submit():
        print(dict(request.form))
        must_be_removed = dict(request.form)
        try:
            del must_be_removed["csrf_token"]
            must_be_removed = [int(x) for x in must_be_removed.keys()]
            TestResult.query.filter(TestResult.id.in_(must_be_removed)).delete()
            db.session.commit()
        except Exception as e:
            if project_name:
                return redirect(url_for("psytest_view_results", project_name=project_name))

            else:
                return "Произошла ошибка", 500
    if project_name:
        return redirect(url_for("psytest_view_results", project_name=project_name))

    else:
        return "Результаты удалены успешно", 200


class LoginForm(FlaskForm):
    password = PasswordField('Password', validators=[InputRequired()])

@app.route('/admin/login', methods=['GET', 'POST'])
def login():
    try:
        form = LoginForm()
        if form.validate_on_submit():
            user = User.query.filter_by(username="admin").first()

            if not user:
                user = User()
                user.username = "admin"
                user.set_password(config["ADMIN_DEFAULT_PASSWORD"])
                db.session.add(user)
                db.session.commit()

            if user is not None and user.check_password(form.password.data):
                flash('Вы вошли как админ')
                login_user(user, remember=True)
                return redirect(url_for('psytest_view_results', project_name=tuple(projects.keys())[0]))
        return render_template('login.html', form=form, logged_in=current_user.is_authenticated)
    except Exception as e:
        add_traceback()
        abort(500, "Произошла ошибка")

@app.route('/admin/logout')
def logout():
    logout_user()
    flash('Вы вышли из аккаунта')
    return redirect(url_for('login'))




def create_excel_result_page(wb: Workbook=None, project_name: str = None):
    # Получаем все результаты тестов для данного проекта
    test_results = TestResult.query.filter(TestResult.project_name == project_name).all()

    # Создаем новую книгу Excel
    if not wb:
        wb = Workbook()
        wb.remove(wb.active)

    # Создаем новую страницу с названием project_name
    sheet = wb.create_sheet(title=project_name)

    # Создаем заголовки для столбцов
    headers = ['Имя', 'IP', 'Время отправки', 'Время прохождения']
    # Добавим заголовки для параметров
    parameter_names = set()
    for test_result in test_results:
        parameter_names.update(parameter.name for parameter in test_result.parameters)
    headers.extend(sorted(parameter_names))

    sheet.freeze_panes = 'A2'

    # Заполняем заголовки в первой строке
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    # Заполняем данные в таблице
    for row, test_result in enumerate(test_results, start=2):
        data = [test_result.name, test_result.ip, test_result.end_time, str(datetime.timedelta(seconds=test_result.duration))]
        data.extend(parameter.value for parameter in sorted(test_result.parameters, key=lambda p: p.name))
        for col, value in enumerate(data, start=1):
            sheet.cell(row=row, column=col, value=value)
            if col == 4:
                sheet.cell(row=row, column=col).number_format = 'HH:MM:SS'

    # Возвращаем страницу
    return wb

def create_excel_results_book():
    # Создаем новую книгу Excel
    wb = Workbook()

    # Удаляем автоматически созданную страницу
    wb.remove(wb.active)

    # Для каждого проекта в словаре создаем новую страницу
    for project_name in projects.keys():
        create_excel_result_page(wb, project_name)

    # Возвращаем книгу
    return wb

@app.route('/api/download', methods=['POST', 'GET'])
def download_results_book():
    # Создаем новую книгу Excel
    wb = create_excel_results_book()

    # Сохраняем данные в поток
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Отправляем файл пользователю
    return send_file(stream, as_attachment=True, download_name='report.xlsx')

@app.route('/api/download/<string:project_name>', methods=['POST', 'GET'])
def download_results_page(project_name):
    # Создаем новую книгу Excel
    wb = create_excel_result_page(project_name=project_name)

    # Сохраняем данные в поток
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    # Отправляем файл пользователю
    return send_file(stream, as_attachment=True, download_name='report.xlsx')

@app.route("/api/update-projects")
def update_projects():
    load_projects()
    return f"<h1>Проекты обновлены {str(projects)}</h1>"

@app.route("/")
def index():
   return "<h1>API работает</h1>"

def create_db_tables():
    with app.app_context():
        db.create_all()

if __name__ == '__main__':
    app.run(host='0.0.0.0')
